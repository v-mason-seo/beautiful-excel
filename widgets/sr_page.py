from pathlib import Path

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QFileDialog,
    QLabel,
    QMessageBox,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

from constants import EXCEL_MAP
from widgets.common import PasteableTable
from widgets.menu_bar import MenuBar


class SRPage(QWidget):
    def __init__(self, excel_loader, app_config, parent=None):
        super().__init__(parent)
        self._loader = excel_loader
        self._app_config = app_config
        self._current_template_path: str | None = None  # 로드된 템플릿 경로
        self._current_category: str = ""

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self._menu_bar = MenuBar()
        layout.addWidget(self._menu_bar)

        self._content = QWidget()
        self._content_layout = QVBoxLayout(self._content)
        self._content_layout.setContentsMargins(16, 16, 16, 16)
        layout.addWidget(self._content)

        self._menu_bar.category_changed.connect(self._on_combo_changed)
        self._menu_bar.save_requested.connect(self._on_save)
        self._menu_bar.settings_requested.connect(self._on_settings)
        self._on_combo_changed(self._menu_bar.current_category())

    # ------------------------------------------------------------------
    # 카테고리 변경 → 템플릿 로드
    # ------------------------------------------------------------------

    def _on_combo_changed(self, text: str):
        self._current_category = text
        filepath = EXCEL_MAP.get(text)
        if not filepath:
            return
        try:
            sheet_data = self._loader.load(filepath)
            self._current_template_path = filepath
            self._build_content(sheet_data, text)
        except Exception as e:
            self._show_error(str(e))

    def _build_content(self, sheet_data: dict, category: str):
        self._clear_content()
        if not sheet_data:
            return

        mapping = self._app_config.get_mapping(category)
        smart_enabled = self._app_config.get("smart_paste_enabled", True)

        if len(sheet_data) == 1:
            name, rows = next(iter(sheet_data.items()))
            table = self._make_table(rows, mapping, smart_enabled)
            self._content_layout.addWidget(table)
        else:
            tabs = QTabWidget()
            for sheet_name, rows in sheet_data.items():
                table = self._make_table(rows, mapping, smart_enabled)
                tabs.addTab(table, sheet_name)
            self._content_layout.addWidget(tabs)

    def _make_table(self, rows: list, mapping: dict, smart_enabled: bool) -> PasteableTable:
        table = PasteableTable()
        table.set_smart_paste_config(smart_enabled, mapping)
        if rows:
            table.load_data(rows[0], rows[1:])
        return table

    # ------------------------------------------------------------------
    # 저장
    # ------------------------------------------------------------------

    def _on_save(self):
        from core.template_exporter import TemplateExporter

        table = self._active_table()
        if table is None:
            QMessageBox.warning(self, "저장 오류", "저장할 테이블을 찾을 수 없습니다.")
            return

        data = table.get_all_data()
        if not any(cell.strip() for row in data for cell in row):
            QMessageBox.warning(self, "저장 오류", "저장할 데이터가 없습니다.")
            return

        default_dir = self._app_config.get("save_path", r"C:\sr_result")
        try:
            Path(default_dir).mkdir(parents=True, exist_ok=True)
        except Exception:
            default_dir = str(Path.home())

        save_path, _ = QFileDialog.getSaveFileName(
            self, "Excel 파일 저장", default_dir,
            "Excel Files (*.xlsx);;All Files (*.*)"
        )
        if not save_path:
            return

        try:
            if self._current_template_path and Path(self._current_template_path).exists():
                data_start = TemplateExporter.detect_data_start_row(self._current_template_path)
                TemplateExporter.save_with_template(
                    template_path=self._current_template_path,
                    data=data,
                    save_path=save_path,
                    data_start_row=data_start,
                )
                mode = "서식 유지 저장 (템플릿 기반)"
            else:
                self._save_plain(data, save_path)
                mode = "일반 저장"

            QMessageBox.information(self, "저장 완료", f"저장되었습니다.\n\n경로: {save_path}\n방식: {mode}")
        except Exception as e:
            QMessageBox.critical(self, "저장 실패", f"저장 중 오류가 발생했습니다:\n{e}")

    def _save_plain(self, data: list, save_path: str):
        """템플릿 없을 때 openpyxl 일반 저장."""
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        for r, row in enumerate(data, start=1):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                if r == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="343A56")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        Path(save_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(save_path)
        wb.close()

    # ------------------------------------------------------------------
    # 설정
    # ------------------------------------------------------------------

    def _on_settings(self):
        from widgets.settings_dialog import SettingsDialog
        dlg = SettingsDialog(self, self._app_config)
        if dlg.exec():
            # 설정 반영: 현재 카테고리 재로드
            self._on_combo_changed(self._current_category)

    # ------------------------------------------------------------------
    # 헬퍼
    # ------------------------------------------------------------------

    def _active_table(self) -> PasteableTable | None:
        """현재 보이는 PasteableTable 반환."""
        if self._content_layout.count() == 0:
            return None
        widget = self._content_layout.itemAt(0).widget()
        if isinstance(widget, PasteableTable):
            return widget
        if isinstance(widget, QTabWidget):
            return widget.currentWidget()
        return None

    def _show_error(self, message: str):
        self._clear_content()
        label = QLabel(f"파일을 불러올 수 없습니다:\n{message}")
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        label.setStyleSheet("color: #888; font-size: 13px; padding: 32px;")
        self._content_layout.addWidget(label)

    def _clear_content(self):
        while self._content_layout.count():
            child = self._content_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
