"""
템플릿 기반 엑셀 저장 - 서식(폰트·병합·테두리)을 유지한 채 데이터만 갱신.

왜 openpyxl인가 (vs win32com):
  - win32com은 DRM 해제 로딩에 사용, 저장은 openpyxl로 서식 보존
  - 개발/CI 환경(비Windows)에서도 동작
  - 병합 셀·페이지 설정은 load_workbook 시 자동 유지
"""

from copy import copy
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


class TemplateExporter:

    @staticmethod
    def save_with_template(
        template_path: str,
        data: List[List[str]],
        save_path: str,
        data_start_row: int = 2,
        data_start_col: int = 1,
    ) -> None:
        """
        템플릿 서식을 유지하면서 data를 써서 save_path에 저장.

        Args:
            template_path : 원본 템플릿 .xlsx 경로
            data          : 헤더 제외 데이터 행 (2D 리스트)
            save_path     : 저장 경로
            data_start_row: 1-base 시작 행 (기본 2 = 헤더 다음 행)
            data_start_col: 1-base 시작 열 (기본 1)
        """
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        template_max_row = ws.max_row

        for row_offset, row_data in enumerate(data):
            current_row = data_start_row + row_offset

            for col_offset, value in enumerate(row_data):
                col = data_start_col + col_offset
                cell = ws.cell(row=current_row, column=col)

                # 템플릿 범위 초과 시 마지막 행 스타일 복사
                if current_row > template_max_row:
                    src = ws.cell(row=template_max_row, column=col)
                    if src.has_style:
                        cell._style = copy(src._style)

                cell.value = value if value != "" else None

        Path(save_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(save_path)
        wb.close()

    @staticmethod
    def detect_data_start_row(template_path: str) -> int:
        """
        템플릿의 헤더 행 수를 감지해 데이터 시작 행(1-base) 반환.
        배경색이 있는 마지막 행 + 1. 감지 실패 시 기본값 2.
        """
        try:
            wb = openpyxl.load_workbook(template_path, read_only=True)
            ws = wb.active
            last_header_row = 1
            for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row)):
                for cell in row:
                    if cell.fill and cell.fill.fgColor:
                        rgb = getattr(cell.fill.fgColor, "rgb", "00000000")
                        if rgb not in ("00000000", "FFFFFFFF", "00FFFFFF"):
                            last_header_row = cell.row
                            break
            wb.close()
            return last_header_row + 1
        except Exception:
            return 2
