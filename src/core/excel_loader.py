"""
엑셀 파일 로더 - XLSX/XLS 파일을 읽어 데이터와 서식 정보를 추출
"""

import openpyxl
from openpyxl.styles import Font, Color
import xlrd
from typing import Dict, List, Tuple, Optional, Any


class ExcelLoader:
    """
    엑셀 파일을 로드하고 데이터 및 서식 정보를 추출하는 클래스
    """

    @staticmethod
    def load_file(file_path: str) -> Dict[str, Any]:
        """
        엑셀 파일을 로드하여 데이터와 서식 정보 반환

        Args:
            file_path: 엑셀 파일 경로

        Returns:
            dict: {
                'data': 2차원 리스트 (행x열),
                'headers': 헤더 리스트,
                'formatting': 서식 정보 딕셔너리
            }

        Raises:
            Exception: 파일 로드 실패시
        """
        try:
            if file_path.endswith('.xlsx'):
                return ExcelLoader._load_xlsx(file_path)
            elif file_path.endswith('.xls'):
                return ExcelLoader._load_xls(file_path)
            else:
                raise ValueError("지원하지 않는 파일 형식입니다. (.xlsx 또는 .xls만 지원)")
        except Exception as e:
            raise Exception(f"엑셀 파일 로드 실패: {str(e)}")

    @staticmethod
    def _load_xlsx(file_path: str) -> Dict[str, Any]:
        """
        XLSX 파일 로드 (openpyxl 사용)

        Args:
            file_path: XLSX 파일 경로

        Returns:
            dict: 데이터와 서식 정보
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # 데이터 추출
        data = []
        formatting = {
            'fonts': {},  # (row, col): {'bold': bool, 'size': int, 'color': str}
            'colors': {},  # (row, col): {'bg_color': str, 'font_color': str}
            'column_widths': {},  # col: width
        }

        # 실제 데이터가 있는 범위 찾기
        max_row = ws.max_row
        max_col = ws.max_column

        # 빈 행/열 제거를 위해 실제 데이터 범위 확인
        actual_max_row = 0
        actual_max_col = 0

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row, col)
                if cell.value is not None and str(cell.value).strip():
                    actual_max_row = max(actual_max_row, row)
                    actual_max_col = max(actual_max_col, col)

        # 데이터가 없으면 빈 결과 반환
        if actual_max_row == 0 or actual_max_col == 0:
            return {
                'data': [],
                'headers': [],
                'formatting': formatting
            }

        # 데이터 읽기
        for row in range(1, actual_max_row + 1):
            row_data = []
            for col in range(1, actual_max_col + 1):
                cell = ws.cell(row, col)
                cell_value = cell.value if cell.value is not None else ""
                row_data.append(str(cell_value))

                # 서식 정보 추출
                if cell.font:
                    formatting['fonts'][(row - 1, col - 1)] = {
                        'bold': cell.font.bold if cell.font.bold is not None else False,
                        'size': cell.font.size if cell.font.size is not None else 11,
                        'name': cell.font.name if cell.font.name else '맑은 고딕',
                        'color': cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
                    }

                # 배경색/글자색 추출
                if cell.fill and hasattr(cell.fill, 'fgColor'):
                    bg_color = None
                    if hasattr(cell.fill.fgColor, 'rgb'):
                        bg_color = cell.fill.fgColor.rgb

                    formatting['colors'][(row - 1, col - 1)] = {
                        'bg_color': bg_color
                    }

            data.append(row_data)

        # 컬럼 너비 정보 추출
        for col in range(1, actual_max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            if col_letter in ws.column_dimensions:
                width = ws.column_dimensions[col_letter].width
                if width:
                    formatting['column_widths'][col - 1] = width

        # 첫 행을 헤더로 사용
        headers = data[0] if data else []

        wb.close()

        return {
            'data': data,
            'headers': headers,
            'formatting': formatting
        }

    @staticmethod
    def _load_xls(file_path: str) -> Dict[str, Any]:
        """
        XLS 파일 로드 (xlrd 사용) - 레거시 지원

        Args:
            file_path: XLS 파일 경로

        Returns:
            dict: 데이터와 서식 정보
        """
        wb = xlrd.open_workbook(file_path, formatting_info=True)
        ws = wb.sheet_by_index(0)

        # 데이터 추출
        data = []
        formatting = {
            'fonts': {},
            'colors': {},
            'column_widths': {}
        }

        # 실제 데이터가 있는 범위 찾기
        actual_max_row = 0
        actual_max_col = 0

        for row in range(ws.nrows):
            for col in range(ws.ncols):
                cell = ws.cell(row, col)
                if cell.value and str(cell.value).strip():
                    actual_max_row = max(actual_max_row, row + 1)
                    actual_max_col = max(actual_max_col, col + 1)

        # 데이터가 없으면 빈 결과 반환
        if actual_max_row == 0 or actual_max_col == 0:
            return {
                'data': [],
                'headers': [],
                'formatting': formatting
            }

        # 데이터 읽기
        for row in range(actual_max_row):
            row_data = []
            for col in range(actual_max_col):
                cell = ws.cell(row, col)
                cell_value = cell.value if cell.value else ""

                # 날짜/시간 형식 처리
                if cell.ctype == xlrd.XL_CELL_DATE:
                    cell_value = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                    cell_value = f"{cell_value[0]}-{cell_value[1]:02d}-{cell_value[2]:02d}"

                row_data.append(str(cell_value))

                # 서식 정보 추출
                try:
                    xf = wb.format_map.get(ws.cell(row, col).xf_index)
                    if xf:
                        font = wb.font_list[xf.font_index] if hasattr(xf, 'font_index') else None
                        if font:
                            formatting['fonts'][(row, col)] = {
                                'bold': font.bold if hasattr(font, 'bold') else False,
                                'size': font.height / 20 if hasattr(font, 'height') else 11,  # twips to points
                                'name': font.name if hasattr(font, 'name') else '맑은 고딕',
                                'color': None
                            }
                except:
                    # 서식 정보 추출 실패시 무시
                    pass

            data.append(row_data)

        # 첫 행을 헤더로 사용
        headers = data[0] if data else []

        return {
            'data': data,
            'headers': headers,
            'formatting': formatting
        }

    @staticmethod
    def extract_column_data(data: List[List[str]], col_index: int, skip_header: bool = True) -> List[str]:
        """
        특정 컬럼의 데이터 추출

        Args:
            data: 2차원 데이터 리스트
            col_index: 컬럼 인덱스
            skip_header: 헤더 제외 여부

        Returns:
            컬럼 데이터 리스트
        """
        start_row = 1 if skip_header else 0
        column_data = []

        for row in range(start_row, len(data)):
            if col_index < len(data[row]):
                column_data.append(data[row][col_index])
            else:
                column_data.append("")

        return column_data
