"""
엑셀 익스포터 - 그리드 데이터를 엑셀 파일로 저장
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional


class ExcelExporter:
    """
    그리드 데이터를 엑셀 파일로 저장하는 클래스
    """

    @staticmethod
    def save_to_excel(
        file_path: str,
        data: List[List[str]],
        headers: Optional[List[str]] = None,
        formatting: Optional[Dict[str, Any]] = None,
        settings: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        데이터를 엑셀 파일로 저장

        Args:
            file_path: 저장할 파일 경로
            data: 2차원 리스트 형태의 데이터
            headers: 헤더 리스트 (선택사항)
            formatting: 서식 정보 딕셔너리 (선택사항)
            settings: 출력 설정 (용지 크기, 방향, 글자 크기 등)

        Raises:
            Exception: 파일 저장 실패시
        """
        try:
            # 새 워크북 생성
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # 데이터가 없으면 빈 파일 저장
            if not data or len(data) == 0:
                wb.save(file_path)
                wb.close()
                return

            # 기본 설정
            default_font_size = 10
            default_font_name = '맑은 고딕'

            if settings:
                default_font_size = settings.get('font_size', 10)

            # 헤더 작성 (첫 행)
            if headers:
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.font = Font(
                        name=default_font_name,
                        size=default_font_size,
                        bold=True
                    )
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center',
                        wrap_text=True
                    )
                    # 헤더 배경색
                    cell.fill = PatternFill(
                        start_color='D9E1F2',
                        end_color='D9E1F2',
                        fill_type='solid'
                    )
                data_start_row = 2
            else:
                data_start_row = 1

            # 데이터 작성
            for row_idx, row_data in enumerate(data, start=data_start_row):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = cell_value

                    # 기본 폰트 적용
                    cell.font = Font(
                        name=default_font_name,
                        size=default_font_size
                    )

                    # 텍스트 정렬
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='center',
                        wrap_text=False
                    )

            # 서식 정보 적용 (있는 경우)
            if formatting:
                ExcelExporter._apply_formatting(ws, formatting, data_start_row)

            # 컬럼 너비 자동 조정
            ExcelExporter._auto_adjust_column_widths(ws, data, headers)

            # 인쇄 설정 (있는 경우)
            if settings:
                ExcelExporter._apply_print_settings(ws, settings)

            # 파일 저장
            wb.save(file_path)
            wb.close()

        except Exception as e:
            raise Exception(f"엑셀 파일 저장 실패: {str(e)}")

    @staticmethod
    def _apply_formatting(ws, formatting: Dict[str, Any], data_start_row: int) -> None:
        """
        서식 정보를 워크시트에 적용

        Args:
            ws: openpyxl 워크시트
            formatting: 서식 정보 딕셔너리
            data_start_row: 데이터 시작 행 (헤더 제외)
        """
        # 폰트 서식 적용
        if 'fonts' in formatting:
            for (row, col), font_info in formatting['fonts'].items():
                cell = ws.cell(row=row + data_start_row, column=col + 1)
                cell.font = Font(
                    name=font_info.get('name', '맑은 고딕'),
                    size=font_info.get('size', 10),
                    bold=font_info.get('bold', False),
                    color=font_info.get('color')
                )

        # 색상 서식 적용
        if 'colors' in formatting:
            for (row, col), color_info in formatting['colors'].items():
                cell = ws.cell(row=row + data_start_row, column=col + 1)
                bg_color = color_info.get('bg_color')
                if bg_color:
                    cell.fill = PatternFill(
                        start_color=bg_color,
                        end_color=bg_color,
                        fill_type='solid'
                    )

        # 컬럼 너비 적용
        if 'column_widths' in formatting:
            for col_idx, width in formatting['column_widths'].items():
                col_letter = get_column_letter(col_idx + 1)
                ws.column_dimensions[col_letter].width = width

    @staticmethod
    def _auto_adjust_column_widths(
        ws,
        data: List[List[str]],
        headers: Optional[List[str]] = None
    ) -> None:
        """
        컬럼 너비를 데이터에 맞게 자동 조정

        Args:
            ws: openpyxl 워크시트
            data: 데이터 리스트
            headers: 헤더 리스트
        """
        # 각 컬럼의 최대 문자 수 계산
        max_lengths = {}

        # 헤더 길이 확인
        if headers:
            for col_idx, header in enumerate(headers):
                max_lengths[col_idx] = len(str(header))

        # 데이터 길이 확인
        for row_data in data:
            for col_idx, cell_value in enumerate(row_data):
                current_length = len(str(cell_value))
                if col_idx not in max_lengths:
                    max_lengths[col_idx] = current_length
                else:
                    max_lengths[col_idx] = max(max_lengths[col_idx], current_length)

        # 컬럼 너비 설정 (최소 8, 최대 50)
        for col_idx, max_length in max_lengths.items():
            col_letter = get_column_letter(col_idx + 1)
            # 한글은 2배, 영문/숫자는 1.2배 여유
            adjusted_width = min(max(max_length * 1.5, 8), 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    @staticmethod
    def _apply_print_settings(ws, settings: Dict[str, Any]) -> None:
        """
        인쇄 설정 적용

        Args:
            ws: openpyxl 워크시트
            settings: 인쇄 설정 딕셔너리
        """
        paper_size = settings.get('paper_size', 'A4')
        orientation = settings.get('orientation', 'landscape')

        # 용지 크기 설정
        if paper_size == 'A4':
            ws.page_setup.paperSize = 9  # A4
        elif paper_size == 'A3':
            ws.page_setup.paperSize = 8  # A3

        # 용지 방향 설정
        if orientation == 'landscape':
            ws.page_setup.orientation = 'landscape'
        else:
            ws.page_setup.orientation = 'portrait'

        # 인쇄 여백 설정 (cm를 inch로 변환: 1cm = 0.3937 inch)
        ws.page_margins.left = 0.394  # 1cm
        ws.page_margins.right = 0.394  # 1cm
        ws.page_margins.top = 0.394  # 1cm
        ws.page_margins.bottom = 0.394  # 1cm

        # 용지에 맞춤
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 높이는 자동

        # 인쇄 제목 (헤더 행 반복)
        if ws.max_row > 1:
            ws.print_title_rows = '1:1'  # 첫 번째 행을 모든 페이지에 반복

    @staticmethod
    def save_with_optimization(
        file_path: str,
        data: List[List[str]],
        headers: Optional[List[str]] = None,
        settings: Optional[Dict[str, Any]] = None,
        optimization_info: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        최적화 정보를 적용하여 엑셀 파일 저장

        Args:
            file_path: 저장할 파일 경로
            data: 2차원 리스트 형태의 데이터
            headers: 헤더 리스트
            settings: 출력 설정
            optimization_info: 최적화 정보 (빈 셀, Bold 처리 등)

        Note:
            optimization_info는 Phase 4에서 구현될 최적화 로직의 결과를 포함
        """
        # 기본 저장 기능 사용
        # Phase 4에서 optimization_info를 활용한 추가 서식 적용 예정
        ExcelExporter.save_to_excel(
            file_path=file_path,
            data=data,
            headers=headers,
            settings=settings
        )
